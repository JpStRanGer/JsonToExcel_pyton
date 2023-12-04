import json
import openpyxl


# Function to read JSON file and output selected information
def read_json_and_output(json_file):
    try:
        with open(json_file, 'r', encoding='utf-8') as json_data:
            data = json.load(json_data)
    except FileNotFoundError:
        print(f"Error: JSON file '{json_file}' not found.")
        return
    except json.JSONDecodeError:
        print(f"Error: Invalid JSON in '{json_file}'.")
        return

    objects = data['objects']
    result = {}

    # loop throug the object list

    # for obj in objects[:3]: # For debugging purpuses. Loopong only first # of objects
    for obj in objects:

        # Store obj.name and obj.id as key and value intu dictionary 'result'
        result.setdefault(obj['name'],obj['id'])

        # loop troug the first set of keys inside the child dic variables inside the object to extract child tags
        for key, variable in obj["variables"].items():

            # Store obj.variables.name and obj.variables.tsref as key and value into dictionary 'result'
            result.setdefault(key,variable['tsref'])
    return result


def write_dict_to_excel(data, excel_file):
    try:
        wb = openpyxl.Workbook()
        sheet = wb.active

        # Add headers to the Excel sheet
        sheet.cell(row=1, column=1, value="TagName")
        sheet.cell(row=1, column=2, value="TsRef")

        # Write dictionary data to the Excel sheet
        for row, (key, value) in enumerate(data.items(), start=2):
            sheet.cell(row=row, column=1, value=key)
            sheet.cell(row=row, column=2, value=value)

        # Save the Excel file
        try:
            wb.save(excel_file)
            print(f"Data has been written to '{excel_file}'")
        except Exception as save_error:
            print(f"Error saving the Excel file: {save_error}")

    except Exception as e:
        print(f"An error occurred: {e}")



# Replace these values with your JSON file and desired keys
json_file = 'AC800.json'
excel_file = "VeasTagTsRef.xlsx"


if len(sys.argv) == 3:
    print(f"you provided {json_file} as json file, and {excel_file} as exel file. generating files....")
    # Extract the two parameters
    json_file = sys.argv[1]
    excel_file = sys.argv[2]
if len(sys.argv) == 2:
    print(f"one parameter provided. Using {sys.argv[1]} as json file, and {excel_file} as exel file.")
    json_file = sys.argv[1]
else:
    print(f"No parameters provided. Using {json_file} as json file, and {excel_file} as exel file.")

print(f"leng of sys.argv: {len(sys.argv)}")

# Call the function to read JSON and output selected information
json_res = read_json_and_output(json_file)
write_dict_to_excel(json_res, excel_file)



